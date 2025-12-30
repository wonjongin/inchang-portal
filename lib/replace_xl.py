#!/usr/bin/env python3

import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import sys

def replace_xl_template(input_file: str, output_file: str, fmt: str):
    workbook = load_workbook(input_file)
    sheet = workbook['Sheet1']

    # fmt 형식: 셀주소::값::셀주소::값...
    # 짝수 인덱스(0, 2, 4, ...): 셀 주소
    # 홀수 인덱스(1, 3, 5, ...): 값 (플레이스홀더 포함)
    fmt_lines = fmt.split('::')
    for i in range(0, len(fmt_lines) - 1, 2):
        cell_address = fmt_lines[i].strip()  # 셀 주소 (예: "B4")
        value_template = fmt_lines[i + 1].strip()  # 값 템플릿 (예: "{{내용}}")
        
        if not cell_address or not value_template:
            continue
    
        cell = sheet[cell_address]
        value = value_template.replace('{{n}}', '\n')
        
        # MergedCell인 경우 처리
        if isinstance(cell, MergedCell):
            # 해당 셀이 포함된 merged range 찾기
            from openpyxl.utils import get_column_letter, column_index_from_string
            col_letter = ''.join([c for c in cell_address if c.isalpha()])
            col_num = column_index_from_string(col_letter)
            row_num = int(''.join([c for c in cell_address if c.isdigit()]))
            
            for merged_range in list(sheet.merged_cells.ranges):
                if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col_num <= merged_range.max_col):
                    # unmerge
                    sheet.unmerge_cells(str(merged_range))
                    # 첫 번째 셀에 값 설정
                    first_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                    first_cell.value = value
                    # 다시 merge
                    sheet.merge_cells(str(merged_range))
                    break
        else:
            # 일반 셀인 경우 직접 값 설정
            cell.value = value
        
    if not os.path.exists(os.path.dirname(output_file)):
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
    workbook.save(output_file)

if __name__ == "__main__":
    output_file = sys.argv[2]
    input_file = sys.argv[1]
    fmt = sys.argv[3]
    replace_xl_template(input_file, output_file, fmt)
    print(f"Replaced {input_file} to {output_file} with fmt {fmt}")